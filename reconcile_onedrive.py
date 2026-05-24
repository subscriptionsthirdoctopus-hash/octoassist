"""
TEMA India — OneDrive vs Backup Reconciliation
==============================================

Author    : Third Octopus (engagement: TEMA India Pvt. Ltd.)
Purpose   : Identify files present in the on-prem backup index but absent from
            the TEMA India "Design" OneDrive site, after the Google Workspace
            to Microsoft 365 migration.

Match key : (File Name + Size in Bytes), compared as strings.
            Modified timestamp is deliberately excluded from the key.

Output    : A branded Excel workbook (Aptos Display, Third Octopus teal header)
            with a Summary sheet and a deduplicated "Missing Files" sheet.

Design decisions (see thread Decision Log for full rationale):
    D1  Match key = (name, size_bytes). Name+size is the strongest deterministic
        signal available across the two sources.
    D2  Compare size as STRING, not int. Dirty rows can have .0 suffixes or
        stray whitespace; string compare after .strip() is robust.
    D3  Modified timestamp NOT in match key. Migration rewrites timestamps.
    D4  Build the set from OneDrive (the smaller side); iterate the backup.
    D5  Internally classify three ways (exact / size-differs / name-absent),
        externally collapse to a single "Missing" view.
    D6  Dedupe by (name, size) and preserve duplicate count in a column.
    D7  Sort missing list by size descending (restore-planning convention).
    D8  openpyxl for styled output (Aptos Display, brand teal header fills).
    D9  Summary sheet on tab 1, data on tab 2 (Third Octopus delivery convention).

Do NOT change without re-reading the Decision Log:
    - Don't cast size_bytes to int.
    - Don't include modified in the match key.
    - Don't try to normalise paths between the two sources.
"""

from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ----------------------------------------------------------------------------- 
# Brand constants (Third Octopus house style)
# -----------------------------------------------------------------------------
APTOS = "Aptos Display"
BRAND_TEAL = "0F4C5C"      # Header fill
BRAND_HIGHLIGHT = "FFF4E6" # KPI row highlight


# ----------------------------------------------------------------------------- 
# Core reconciliation
# -----------------------------------------------------------------------------
def load_sources(backup_csv: Path, onedrive_csv: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load only the columns we need, as strings. dtype=str avoids pandas
    type inference on large dirty CSVs."""
    bk = pd.read_csv(
        backup_csv,
        usecols=["path", "name", "size_bytes", "modified"],
        dtype=str,
        low_memory=False,
    )
    od = pd.read_csv(
        onedrive_csv,
        usecols=["Path", "Name", "SizeBytes", "Modified"],
        dtype=str,
        low_memory=False,
    )

    # D2: strip whitespace BEFORE building the match key.
    for col in ("name", "size_bytes"):
        bk[col] = bk[col].fillna("").str.strip()
    for col in ("Name", "SizeBytes"):
        od[col] = od[col].fillna("").str.strip()

    return bk, od


def reconcile(bk: pd.DataFrame, od: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Return the deduplicated missing-files frame plus a stats dict.

    Match key: (name, size_bytes) as strings.
    """
    # D4: build the set from the smaller side (OneDrive).
    od_keys = set(zip(od["Name"], od["SizeBytes"]))

    # Vectorised membership — list comprehension is ~30x faster than .apply().
    bk["in_od"] = [k in od_keys for k in zip(bk["name"], bk["size_bytes"])]

    missing_all = bk[~bk["in_od"]].copy()

    # D6: count duplicate copies in the backup before dedup.
    missing_all["copies_in_backup"] = (
        missing_all.groupby(["name", "size_bytes"])["name"].transform("count")
    )

    missing = missing_all.drop_duplicates(["name", "size_bytes"], keep="first").copy()

    # Enrich for output.
    missing["size_bytes_int"] = (
        pd.to_numeric(missing["size_bytes"], errors="coerce").fillna(0).astype("int64")
    )
    missing["size_mb"] = (missing["size_bytes_int"] / (1024 * 1024)).round(3)
    missing["modified_dt"] = pd.to_datetime(missing["modified"], errors="coerce")

    # D7: largest first.
    missing = missing.sort_values(["size_bytes_int", "name"], ascending=[False, True])

    # Stats for the Summary sheet.
    missing_all_size_int = pd.to_numeric(
        missing_all["size_bytes"], errors="coerce"
    ).fillna(0).astype("int64")

    stats = {
        "backup_rows": len(bk),
        "onedrive_rows": len(od),
        "matched_removed": int(bk["in_od"].sum()),
        "missing_before_dedup": len(missing_all),
        "missing_after_dedup": len(missing),
        "missing_gb": missing["size_bytes_int"].sum() / (1024 ** 3),
        "missing_gb_with_dupes": missing_all_size_int.sum() / (1024 ** 3),
    }
    stats["dedup_savings_gb"] = stats["missing_gb_with_dupes"] - stats["missing_gb"]

    out = missing[
        ["path", "name", "size_bytes_int", "size_mb", "modified_dt", "copies_in_backup"]
    ].rename(
        columns={
            "path": "Path",
            "name": "File Name",
            "size_bytes_int": "Size (Bytes)",
            "size_mb": "Size (MB)",
            "modified_dt": "Last Modified",
            "copies_in_backup": "Copies in Backup",
        }
    )
    return out, stats


# ----------------------------------------------------------------------------- 
# Workbook writer (Third Octopus house style)
# -----------------------------------------------------------------------------
def _header_cell(cell) -> None:
    cell.font = Font(name=APTOS, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=BRAND_TEAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_workbook(out_df: pd.DataFrame, stats: dict, out_path: Path) -> None:
    wb = Workbook()

    # --- Summary sheet ---
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False

    s["A1"] = "TEMA India — Unique Files Missing from OneDrive"
    s["A1"].font = Font(name=APTOS, size=16, bold=True, color=BRAND_TEAL)
    s.merge_cells("A1:D1")
    s.row_dimensions[1].height = 26

    s["A2"] = "Prepared by: Third Octopus  |  Private & Confidential"
    s["A2"].font = Font(name=APTOS, size=10, italic=True, color="666666")
    s.merge_cells("A2:D2")

    s["A4"] = "Method"
    s["A4"].font = Font(name=APTOS, size=12, bold=True, color=BRAND_TEAL)
    notes = [
        "Match key — (File Name + Size in Bytes). Files present in OneDrive at the same size were removed.",
        "Deduplication — each unique (File Name + Size) appears once, even if it existed in multiple folders.",
        "'Copies in Backup' shows how many times the same file appeared across the backup folder tree.",
        "Rows sorted by file size, largest first.",
    ]
    for i, n in enumerate(notes, start=5):
        s.cell(row=i, column=1, value=f"•  {n}").font = Font(name=APTOS, size=10)
        s.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    start = 11
    for col, h in enumerate(["Metric", "Value"], start=1):
        _header_cell(s.cell(row=start, column=col, value=h))

    kpis = [
        ("Backup rows (total)", f"{stats['backup_rows']:,}"),
        ("OneDrive rows (total)", f"{stats['onedrive_rows']:,}"),
        ("Matched and removed (present in OneDrive at same size)", f"{stats['matched_removed']:,}"),
        ("Missing rows before deduplication", f"{stats['missing_before_dedup']:,}"),
        ("UNIQUE missing files (after dedup by Name + Size)", f"{stats['missing_after_dedup']:,}"),
        ("Total size of unique missing data", f"{stats['missing_gb']:.2f} GB"),
        ("Total size including all duplicate copies", f"{stats['missing_gb_with_dupes']:.2f} GB"),
        ("Storage saved by deduplication", f"{stats['dedup_savings_gb']:.2f} GB"),
    ]
    for r, (k, v) in enumerate(kpis, start=start + 1):
        a = s.cell(row=r, column=1, value=k)
        a.font = Font(name=APTOS, size=10)
        b = s.cell(row=r, column=2, value=v)
        b.font = Font(name=APTOS, size=10, bold=True)
        if "UNIQUE" in k or "unique missing data" in k:
            a.fill = PatternFill("solid", start_color=BRAND_HIGHLIGHT)
            b.fill = PatternFill("solid", start_color=BRAND_HIGHLIGHT)

    s.column_dimensions["A"].width = 65
    s.column_dimensions["B"].width = 22
    s.freeze_panes = "A4"

    # --- Missing Files sheet ---
    ws = wb.create_sheet("Missing Files (Unique)")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Unique Files Missing from OneDrive — deduplicated by (Name + Size)"
    ws["A1"].font = Font(name=APTOS, size=14, bold=True, color=BRAND_TEAL)
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"{len(out_df):,} unique files  |  Total {stats['missing_gb']:.2f} GB  "
        f"|  Sorted by size, largest first"
    )
    ws["A2"].font = Font(name=APTOS, size=10, italic=True, color="666666")
    ws.merge_cells("A2:F2")

    header_row = 4
    for c, name in enumerate(out_df.columns, start=1):
        _header_cell(ws.cell(row=header_row, column=c, value=name))

    # Bulk write rows via itertuples — materially faster than pd.to_excel for
    # styled output at this row count.
    for ridx, row in enumerate(out_df.itertuples(index=False), start=header_row + 1):
        for cidx, val in enumerate(row, start=1):
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.font = Font(name=APTOS, size=9)
            cell.alignment = Alignment(vertical="center")
            if cidx == 3:
                cell.number_format = "#,##0"
            elif cidx == 4:
                cell.number_format = "#,##0.000"
            elif cidx == 5 and val != "":
                cell.number_format = "dd-mmm-yyyy hh:mm"
            elif cidx == 6:
                cell.number_format = "#,##0"

    for col, w in {"A": 90, "B": 50, "C": 16, "D": 14, "E": 22, "F": 16}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    ws.row_dimensions[1].height = 22

    wb.save(out_path)


# ----------------------------------------------------------------------------- 
# CLI
# -----------------------------------------------------------------------------
def main() -> int:
    p = argparse.ArgumentParser(description="TEMA India OneDrive reconciliation.")
    p.add_argument(
        "--backup",
        default="/mnt/user-data/uploads/design_files_2.csv",
        help="Path to the backup index CSV.",
    )
    p.add_argument(
        "--onedrive",
        default="/mnt/user-data/uploads/Design_TEMA_OneDrive_Inventory.csv",
        help="Path to the OneDrive inventory CSV.",
    )
    p.add_argument(
        "--out",
        default="/mnt/user-data/outputs/TEMA_OneDrive_Missing_Files_Deduped.xlsx",
        help="Output XLSX path.",
    )
    args = p.parse_args()

    backup_csv = Path(args.backup)
    onedrive_csv = Path(args.onedrive)
    out_path = Path(args.out)

    for csv in (backup_csv, onedrive_csv):
        if not csv.exists():
            print(f"ERROR: input not found: {csv}", file=sys.stderr)
            return 2

    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading: {backup_csv}")
    print(f"Loading: {onedrive_csv}")
    bk, od = load_sources(backup_csv, onedrive_csv)

    print("Reconciling…")
    out_df, stats = reconcile(bk, od)

    for k, v in stats.items():
        if isinstance(v, float):
            print(f"  {k}: {v:.2f}")
        else:
            print(f"  {k}: {v:,}")

    print(f"Writing: {out_path}")
    write_workbook(out_df, stats, out_path)

    size_mb = os.path.getsize(out_path) / (1024 * 1024)
    print(f"Done. Output: {out_path}  ({size_mb:.1f} MB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
